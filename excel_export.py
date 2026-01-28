from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart
from typing import Dict, Any, List, Optional
from datetime import datetime

def create_excel_report(
    filepath: str,
    project_title: str,
    inputs: Dict[str, Any],
    outputs: Dict[str, Any],
    warnings: List[str],
    usd_try_rate: Optional[float],
    rate_source: Optional[str]
):
    """Create comprehensive Excel report with charts"""
    
    wb = Workbook()
    
    # ============================================================================
    # SHEET 1: SUMMARY
    # ============================================================================
    ws_summary = wb.active
    ws_summary.title = "Ozet"
    
    # Header
    ws_summary['A1'] = "KONUT PROJESI FIZIBILITE RAPORU"
    ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.merge_cells('A1:D1')
    ws_summary.row_dimensions[1].height = 30
    
    # Project info
    ws_summary['A3'] = "Proje:"
    ws_summary['B3'] = project_title
    ws_summary['B3'].font = Font(bold=True)
    
    ws_summary['A4'] = "Tarih:"
    ws_summary['B4'] = datetime.now().strftime('%d.%m.%Y')
    
    if usd_try_rate:
        ws_summary['A5'] = "Kur (USD/TRY):"
        ws_summary['B5'] = f"{usd_try_rate:.4f}"
        ws_summary['C5'] = rate_source or ""
    
    ws_summary['A6'] = "Hazirlayan:"
    ws_summary['B6'] = "Dr. Omur Tezcan / GGtech"
    
    # Key metrics header
    ws_summary['A8'] = "ANA METRIKLER"
    ws_summary['A8'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_summary['A8'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_summary.merge_cells('A8:D8')
    
    # Metrics table
    metrics = [
        ["Metrik", "Deger", "Birim", "Detay"],
        ["Satilabilir Alan", outputs.get('satilabilir_alan_m2', 0), "m²", ""],
        ["Toplam Insaat Alani", outputs.get('toplam_insaat_alani_m2', 0), "m²", "Otopark dahil"],
        ["Konut Adedi", int(outputs.get('yaklasik_konut_adedi', 0)), "adet", ""],
        ["Toplam Maliyet (USD)", outputs.get('toplam_proje_maliyeti_usd', 0), "$", ""],
        ["Toplam Maliyet (TL)", outputs.get('toplam_proje_maliyeti_try', 0), "₺", ""],
        ["Basabas Fiyat (USD)", outputs.get('breakeven_usd_m2', 0), "$/m²", ""],
        ["Basabas Fiyat (TL)", outputs.get('breakeven_try_m2', 0), "₺/m²", ""],
    ]
    
    row = 9
    for metric in metrics:
        for col, value in enumerate(metric, start=1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if row == 9:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            if col == 2 and row > 9:  # Number formatting
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
        row += 1
    
    # Pricing strategy
    ws_summary['A18'] = "SATIS FIYAT STRATEJISI"
    ws_summary['A18'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_summary['A18'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_summary.merge_cells('A18:D18')
    
    pricing = [
        ["Hedef", "USD/m²", "TL/m²", "Aciklama"],
        ["Basabas", outputs.get('breakeven_usd_m2', 0), outputs.get('breakeven_try_m2', 0), "Maliyet karsilama"],
        ["%10 Kar", outputs.get('target_10_usd_m2', 0), outputs.get('target_10_try_m2', 0), "Muhafazakar"],
        ["%30 Kar", outputs.get('target_30_usd_m2', 0), outputs.get('target_30_try_m2', 0), "Dengeli"],
        ["%50 Kar", outputs.get('target_50_usd_m2', 0), outputs.get('target_50_try_m2', 0), "Agresif"],
    ]
    
    row = 19
    for price in pricing:
        for col, value in enumerate(price, start=1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if row == 19:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            if col in [2, 3] and row > 19:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
        row += 1
    
    # Revenue section (if sales price exists)
    if outputs.get("satis_birim_fiyat_usd_m2"):
        ws_summary['A25'] = "GELIR & KARLILIK"
        ws_summary['A25'].font = Font(size=14, bold=True, color="FFFFFF")
        ws_summary['A25'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        ws_summary.merge_cells('A25:D25')
        
        revenue = [
            ["Metrik", "USD", "TL", "Oran"],
            ["Satis Fiyati", outputs.get('satis_birim_fiyat_usd_m2', 0), outputs.get('satis_birim_fiyat_try_m2', 0), ""],
            ["Hasilat", outputs.get('proje_hasilati_usd', 0), outputs.get('proje_hasilati_try', 0), ""],
            ["Kar", outputs.get('proje_kari_usd', 0), outputs.get('proje_kari_try', 0), ""],
            ["Brut Karlilik", "", "", outputs.get('brut_karlilik_orani', 0)],
        ]
        
        row = 26
        for rev in revenue:
            for col, value in enumerate(rev, start=1):
                cell = ws_summary.cell(row=row, column=col, value=value)
                if row == 26:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                if col in [2, 3] and row > 26:
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
                if col == 4 and row == 30:  # Profit margin
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.0%'
            row += 1
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 25
    
    # ============================================================================
    # SHEET 2: DETAILED INPUTS
    # ============================================================================
    ws_inputs = wb.create_sheet("Girdiler")
    
    ws_inputs['A1'] = "PROJE GIRDI PARAMETRELERI"
    ws_inputs['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_inputs['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    ws_inputs.merge_cells('A1:C1')
    
    input_labels = {
        "arsa_alani_m2": "Arsa Alani (m²)",
        "emsal": "Emsal",
        "satilabilir_katsayi": "Satilabilir Alan Katsayisi",
        "otopark_tipi": "Otopark Tipi",
        "otopark_katsayi": "Otopark Katsayisi",
        "konut_sinifi": "Konut Sinifi",
        "insaat_maliyet_usd_m2": "Insaat Maliyeti ($/m²)",
        "arsa_toplam_degeri_usd": "Arsa Degeri ($)",
        "ortalama_konut_m2": "Ortalama Konut (m²)",
        "satis_birim_fiyat_usd_m2": "Satis Fiyati ($/m²)",
    }
    
    row = 3
    ws_inputs.cell(row=row, column=1, value="Parametre").font = Font(bold=True)
    ws_inputs.cell(row=row, column=2, value="Deger").font = Font(bold=True)
    ws_inputs.cell(row=row, column=3, value="Not").font = Font(bold=True)
    row += 1
    
    for key, label in input_labels.items():
        if key in inputs:
            ws_inputs.cell(row=row, column=1, value=label)
            ws_inputs.cell(row=row, column=2, value=inputs[key])
            row += 1
    
    ws_inputs.column_dimensions['A'].width = 30
    ws_inputs.column_dimensions['B'].width = 20
    ws_inputs.column_dimensions['C'].width = 30
    
    # ============================================================================
    # SHEET 3: COST BREAKDOWN with PIE CHART
    # ============================================================================
    ws_cost = wb.create_sheet("Maliyet Dagilimi")
    
    ws_cost['A1'] = "MALIYET DAGILIMI ANALIZI"
    ws_cost['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_cost['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    ws_cost.merge_cells('A1:C1')
    
    # Cost breakdown data
    cost_data = [
        ["Maliyet Kalemi", "USD", "Oran"],
        ["Arsa Degeri", outputs.get('arsa_degeri_usd', 0), ""],
        ["Insaat Maliyeti", outputs.get('insaat_maliyeti_usd', 0), ""],
        ["TOPLAM", outputs.get('toplam_proje_maliyeti_usd', 0), ""],
    ]
    
    # Calculate percentages
    total = outputs.get('toplam_proje_maliyeti_usd', 1)
    if total > 0:
        cost_data[1][2] = outputs.get('arsa_degeri_usd', 0) / total
        cost_data[2][2] = outputs.get('insaat_maliyeti_usd', 0) / total
        cost_data[3][2] = 1.0
    
    row = 3
    for cost in cost_data:
        for col, value in enumerate(cost, start=1):
            cell = ws_cost.cell(row=row, column=col, value=value)
            if row == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            if col == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
            if col == 3 and isinstance(value, (int, float)):
                cell.number_format = '0.0%'
        row += 1
    
    # Add pie chart
    pie = PieChart()
    labels = Reference(ws_cost, min_col=1, min_row=4, max_row=5)
    data = Reference(ws_cost, min_col=2, min_row=3, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Maliyet Dagilimi"
    ws_cost.add_chart(pie, "E3")
    
    ws_cost.column_dimensions['A'].width = 25
    ws_cost.column_dimensions['B'].width = 20
    ws_cost.column_dimensions['C'].width = 15
    
    # ============================================================================
    # SHEET 4: WARNINGS
    # ============================================================================
    if warnings:
        ws_warn = wb.create_sheet("Uyarilar")
        
        ws_warn['A1'] = "UYARILAR VE NOTLAR"
        ws_warn['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws_warn['A1'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        ws_warn.merge_cells('A1:B1')
        
        row = 3
        for i, warning in enumerate(warnings, start=1):
            ws_warn.cell(row=row, column=1, value=i)
            ws_warn.cell(row=row, column=2, value=warning)
            row += 1
        
        ws_warn.column_dimensions['A'].width = 5
        ws_warn.column_dimensions['B'].width = 80
    
    # Save
    wb.save(filepath)

def create_comparison_excel(
    filepath: str,
    scenarios: List[Dict[str, Any]]
):
    """Create Excel with multiple scenario comparison"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Senaryo Karsilastirmasi"
    
    # Header
    ws['A1'] = "SENARYO KARSILASTIRMA ANALIZI"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    ws.merge_cells(f'A1:{chr(65 + len(scenarios))}1')
    
    # Column headers
    ws['A3'] = "Metrik"
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    
    for i, scenario in enumerate(scenarios):
        col = chr(66 + i)  # B, C, D...
        ws[f'{col}3'] = scenario.get('name', f'Senaryo {i+1}')
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    
    # Metrics to compare
    comparison_metrics = [
        ("Arsa Alani (m²)", "arsa_alani_m2", "input"),
        ("Emsal", "emsal", "input"),
        ("Satis Fiyati ($/m²)", "satis_birim_fiyat_usd_m2", "input"),
        ("", "", ""),  # Empty row
        ("Satilabilir Alan (m²)", "satilabilir_alan_m2", "output"),
        ("Konut Adedi", "yaklasik_konut_adedi", "output"),
        ("Toplam Maliyet ($)", "toplam_proje_maliyeti_usd", "output"),
        ("Basabas ($/m²)", "breakeven_usd_m2", "output"),
        ("Hasilat ($)", "proje_hasilati_usd", "output"),
        ("Kar ($)", "proje_kari_usd", "output"),
        ("Brut Karlilik", "brut_karlilik_orani", "output"),
    ]
    
    row = 4
    for label, key, source in comparison_metrics:
        ws[f'A{row}'] = label
        if label:  # Not empty row
            ws[f'A{row}'].font = Font(bold=True if not source else False)
        
        for i, scenario in enumerate(scenarios):
            col = chr(66 + i)
            if source == "input":
                value = scenario.get('inputs', {}).get(key, "")
            elif source == "output":
                value = scenario.get('outputs', {}).get(key, "")
            else:
                value = ""
            
            cell = ws[f'{col}{row}']
            cell.value = value
            
            # Formatting
            if key == "brut_karlilik_orani" and isinstance(value, (int, float)):
                cell.number_format = '0.0%'
            elif isinstance(value, (int, float)) and value > 100:
                cell.number_format = '#,##0'
        
        row += 1
    
    # Add comparison chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Kar Karsilastirmasi"
    chart.x_axis.title = "Senaryo"
    chart.y_axis.title = "Kar (USD)"
    
    # Find profit row
    profit_row = None
    for i, (label, key, _) in enumerate(comparison_metrics, start=4):
        if key == "proje_kari_usd":
            profit_row = i
            break
    
    if profit_row:
        data = Reference(ws, min_col=2, max_col=1+len(scenarios), min_row=profit_row, max_row=profit_row)
        cats = Reference(ws, min_col=2, max_col=1+len(scenarios), min_row=3, max_row=3)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{row+2}")
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    for i in range(len(scenarios)):
        col = chr(66 + i)
        ws.column_dimensions[col].width = 20
    
    wb.save(filepath)
